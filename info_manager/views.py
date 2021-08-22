import sqlite3,os,datetime
import xlrd
from io import BytesIO
from django.db.models import Q
from django.shortcuts import render,redirect
from django.http import HttpResponse
from .models import InfoPost_7,InfoPost_8,InfoPost_9
from .froms import InfoPostForm_7,InfoPostForm_8,InfoPostForm_9
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from openpyxl.workbook import Workbook
from django.utils.http import urlquote
# Create your views here.

# 试题列表
@login_required(login_url='/userprofile/login/')
def info_list(request,index):
    search = request.GET.get('search')
    obj_list = [InfoPost_7,InfoPost_8,InfoPost_9]
    obj = obj_list[index]

    if search:
        qes_infos = obj.objects.filter(
            Q(uploader_id=request.user.id) &
            Q(question__icontains=search) |
            Q(question_info__icontains=search))
    elif request.user.id == 1:
        qes_infos = obj.objects.all()
    else:
        # search=''
        qes_infos = obj.objects.filter(uploader_id=request.user.id)

    paginator = Paginator(qes_infos,5)
    page = request.GET.get('page')
    qes_info = paginator.get_page(page)
    context = {'qes_info':qes_info,'search':search,'index':index}
    return render(request,'info_manager/info_list.html',context)

@login_required(login_url='/userprofile/login/')
@csrf_exempt # 要求不验证csrf_token
# 上传函数
def upload_index(request,grade):
    if request.method == "POST":
        file_obj = request.FILES.get("upload_file")
        if file_obj.name.split('.')[-1] == 'xls':
            with open(settings.UPLOAD_ROOT+'/' + file_obj.name, "wb") as f:
                for chunk in file_obj.chunks():
                    f.write(chunk)
                    handle_file(file_obj.name,grade)
                # return HttpResponse('上传成功')
            return redirect("info_manager:upload_finished")
        else:
            return HttpResponse("您上传的文件格式不是xls，请转换您的文件格式")
    return render(request, "info_manager/upload_index.html")

# 处理文件
def handle_file(file,post_num):
    # 打开excel文件
    read_file = xlrd.open_workbook(settings.UPLOAD_ROOT + '/' + file)
    sheet = read_file.sheet_by_index(0)
    # 获取excel的行与列
    n_rows = sheet.nrows
    n_cols = sheet.ncols
    # 编写sql
    sql = "INSERT INTO INFO_MANAGER_INFOPOST_{} (title,question,option_a,option_b,option_c,option_d,answer,question_id," \
          "question_info,created,uploader_id) VALUES (?,?,?,?,?,?,?,?,?,?,?)".format(post_num)
    list = []
    for i in range(1,n_rows):
        row = sheet.row_values(i)
        # test
        row[-2] = datetime.datetime.now()
        row = tuple(row)
        list.append(row)

    # 连接数据库
    conn = sqlite3.connect('db.sqlite3')
    cur = conn.cursor()
    cur.executemany(sql,list)
    # 提交
    conn.commit()
    # 关闭数据库
    conn.close()

def upload_finished(request):
    return render(request,'info_manager/upload_finished.html')

# 修改函数
@login_required(login_url='/userprofile/login/')
def info_update(request, grade, id):
    # 获取需要修改的具体文章对象
    obj_list = [InfoPost_7,InfoPost_8,InfoPost_9]
    form_list = [InfoPostForm_7, InfoPostForm_8, InfoPostForm_9]

    obj = obj_list[grade]
    info = obj.objects.get(id=id)
    # 判断用户是否为 POST 提交表单数据
    if request.method == "POST":
        # 将提交的数据赋值到表单实例中

        info_post_form = form_list[grade](data=request.POST)
        # 判断提交的数据是否满足模型的要求
        if info_post_form.is_valid():
            # 保存新写入的 title、body 数据并保存
            info.title = request.POST['title']
            info.question = request.POST['question']
            info.option_a = request.POST['option_a']
            info.option_b = request.POST['option_b']
            info.option_c = request.POST['option_c']
            info.option_d = request.POST['option_d']
            info.answer = request.POST['answer']
            info.question_id = request.POST['question_id']
            info.question_info = request.POST['question_info']
            info.save()
            # 完成后返回到修改后的文章中。需传入文章的 id 值
            # return redirect("info_manager:info_index")
            return info_list(request,index=grade)
        # 如果数据不合法，返回错误信息
        else:
            return HttpResponse("表单内容有误，请重新填写。")

    # 如果用户 GET 请求获取数据
    else:
        # 创建表单类实例
        info_post_form = form_list[grade]()
        # 赋值上下文，将 article 文章对象也传递进去，以便提取旧的内容
        context = { 'info': info, 'info_post_form': info_post_form }
        # 将响应返回到模板中
        return render(request, 'info_manager/info_update.html', context)

@login_required(login_url='/userprofile/login/')
def info_index(request):
    return render(request,'info_manager/info_index.html')

@login_required(login_url='userprofile/login/')
def export_excel(request,index):
    wb = Workbook()  # 生成一个工作簿（即一个Excel文件）
    wb.encoding = 'utf-8'
    sheet1 = wb.active  # 获取第一个工作表（sheet1）
    row_one = ['qes','a','b','c','d','ans']
    for i in range(1, len(row_one) + 1):  # 从第一行开始写，因为Excel文件的行号是从1开始，列号也是从1开始
        # 从row=1，column=1开始写，即将row_one的数据依次写入第一行
        sheet1.cell(row=1, column=i).value = row_one[i - 1]
    obj_list = [InfoPost_7,InfoPost_8,InfoPost_9]
    mod = obj_list[index]
    all_obj = mod.objects.all()
    for obj in all_obj:
        max_row = sheet1.max_row + 1  # 获取到工作表的最大行数并加1
        obj_info = [obj.question,obj.option_a,obj.option_b,obj.option_c,obj.option_d,obj.answer]
        for x in range(1, len(obj_info) + 1):  # 将每一个对象的所有字段的信息写入一行内
            sheet1.cell(row=max_row, column=x).value = obj_info[x - 1]

    # 准备写入到IO中
    output = BytesIO()
    wb.save(output)  # 将Excel文件内容保存到IO中
    output.seek(0)  # 重新定位到开始
    # 设置HttpResponse的类型
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
    ctime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    file_name = 'english_grade_{}_{}.xls'.format(index,ctime)  # 给文件名中添加日期时间
    file_name = urlquote(file_name)  # 使用urlquote()方法解决中文无法使用的问题
    response['Content-Disposition'] = 'attachment; filename=%s' % file_name
    # response.write(output.getvalue())	 # 在设置HttpResponse的类型时，如果给了值，可以不写这句
    return response